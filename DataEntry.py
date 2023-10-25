import os
import sys
import tkinter as tk
import openpyxl as pyxl
#Importing ttk from tkinter allows for more widgets to be used (e.g. comboboxes, notebooks and progressbars)
from tkinter import messagebox, ttk, filedialog
from openpyxl import workbook

os.system("cls")

"""
Example: Data Entry Form

Note: Use paint to plan the layout so you know where each grid/frame needs to be initialised
"""
class dataEntry:
    """
    Window Initialisation
    """
    def __init__(data):
        """
        0.0 Window Initialisation [GUI WINDOW]
        """
        data.root = tk.Tk()
        #data.root.geometry('500x500')
        data.root.title("Data Entry Form")
        
        data.style = ttk.Style(data.root)
        data.style.theme_use('default')
        
        data.filepath=None
        
        """
        0.0 Toolbar
        """
        data.menuBar = tk.Menu(data.root)
        data.fileMenu()
        data.helpMenu()   
        data.root.config(menu=data.menuBar)   
        

        
        """
        1.0 User Information [FRAME]
        """
        data.rootFrame = data.create_widget(widgetType = 'frame', layout='pack', frame = data.root)
        data.userInfo = data.create_widget(widgetType='labelframe', layout='grid', frame=data.rootFrame, labelText="User Information",
                                           row=0, col=0, xPad=10, yPad=10, stickyVal=tk.NSEW)
        """
        1.1 Forename [ENTRY]
        """
        data.forenameLabel = data.create_widget(widgetType='label', layout='grid', frame=data.userInfo, labelText="Forename", row=0, col=0)
        data.forenameEntry = data.create_widget(widgetType='entry', layout='grid', frame=data.userInfo, entryWidth=25, row=1, col=0)
        
        """
        1.2 Surname [ENTRY]
        """
        data.surnameLabel = data.create_widget(widgetType='label', layout='grid', frame=data.userInfo, labelText="Surname", row=0, col=1)
        data.surnameEntry = data.create_widget(widgetType='entry', layout='grid', frame=data.userInfo, entryWidth=25, row=1, col=1)

        """
        1.3 Title [COMBOBOX]
        """      
        data.titleList = ["", "Mr.", "Mrs.", "Ms.", "Mx.", "Dr.", "Other", "Prefer not to say" ]
        data.titleCombo = data.valid_combobox(frame=data.userInfo, labelText='Title', row=0, col=2, list=data.titleList, width=12,
                                              valType='title')
        
        """
        1.4 Age [SPINBOX]
        """
        data.ageMin, data.ageMax = 18, 110
        data.ageSpinbox = data.valid_spinbox(row=2, col=0, min=data.ageMin, max=data.ageMax, width=5,
                                             labelText="Age", frame= data.userInfo, valType='age')
        """
        1.5 Nationality [COMBOBOX]
        """
        data.natList = ["", "Asian/Asian British", "Black/African/Caribbean/Black British", "White", "Mixed/Multiple", "Other", "Prefer not to say"]
        data.natCombo = data.valid_combobox(frame=data.userInfo, labelText='Nationality', row=2, col=1, list=data.natList, width=12,
                                              valType='nat')

        """
        2.0 Courses [LABELFRAME]
        """
        data.coursesFrame = data.create_widget(widgetType='labelframe', layout='grid', frame=data.rootFrame, labelText="Courses",
                                               row=1, col=0, xPad=10, yPad=10, stickyVal=tk.NSEW)
        """
        2.1 Registration Status [TICKBOX]
        """
        data.regLabel = data.create_widget(widgetType='label', layout='grid', frame=data.coursesFrame, labelText="Registration Status", row=0, col=0)
        data.regTick, data.checkReg = data.create_widget(widgetType='tick', layout='grid', frame=data.coursesFrame,
                                                         labelText="Currently Registered?", row=1, col=0)
        
        """
        2.2 No. Courses [SPINBOX]
        """
        data.coursesMin, data.coursesMax = 0, 10
        data.coursesSpinbox = data.valid_spinbox(row=0, col=1, min=data.coursesMin, max=data.coursesMax, width=5,
                                                 labelText="No. Completed Courses", frame= data.coursesFrame, valType='courses')
        
        """
        2.3 No. Semesters [SPINBOX]
        """
        data.semestersMin, data.semestersMax= 0, 16
        data.semestersSpinbox = data.valid_spinbox(row=0, col=2, min=data.semestersMin, max=data.semestersMax, width=5,
                                                   labelText="No. Completed Semesters", frame= data.coursesFrame, valType='semesters')
        
        """
        3.0 Terms and Conditions [LABELFRAME]
        """
        data.termsFrame = data.create_widget(widgetType='labelframe', layout='grid', frame=data.rootFrame, labelText="Terms & Conditions",
                                             row=2, col=0, xPad=10, yPad=10, stickyVal=tk.NSEW)
        
        """
        3.1 Terms and Conditions Checkbox [TICKBOX]
        """
        data.termsTick, data.checkTerms = data.create_widget(widgetType='tick', layout='grid', frame=data.termsFrame,
                                                             labelText="I accept the", row=0, col=0)
        data.termsLink = data.create_widget(widgetType='label', layout='grid', frame=data.termsFrame,
                                            labelText="terms and conditions", stickyVal=tk.NSEW,row=0, col=1)
        data.termsLink.configure(font='default 10 underline',fg="blue", cursor="hand2")
        data.termsLink.bind("<Button-1>", lambda e: data.open_terms())
        
        """
        4.0 "Enter Data" Button [BUTTON]
        """
        data.enterButton = data.create_widget(widgetType='button', layout='grid',frame=data.rootFrame, labelText="Enter Data",
                                              row=3, col=0, stickyVal=tk.NSEW, xPad=20,yPad=10, cmd=data.enter_data)
        
        """
        5.0 "Clear" Button [BUTTON]       
        """
        data.clearButton = data.create_widget(widgetType='button', layout='grid',frame=data.rootFrame, labelText="Clear",
                                              row=4, col=0, stickyVal=tk.NSEW, xPad=20, yPad=5, cmd= lambda: data.clear(1))
        """
        X.X Setting Pad Values
        """
        data.pad_grid(data.userInfo, 5, 3)
        data.pad_grid(data.coursesFrame, 5, 3) 
        data.pad_grid(data.termsFrame,0, 3)       

        #KeyBinds        
        data.root.bind_all("<Button-1>", lambda e: data.focus(e))
        data.root.bind_all("<Return>", data.return_check)
        data.root.bind_all("<F10>", data.auto_fill)
        
        #adds a quit option to the exit button
        data.root.protocol("WM_DELETE_WINDOW", lambda: data.quit(1))

        #ConstructWidgets
        tk.mainloop()
    
    #FUNCTIONS    
    """
    WIDGET CREATION
    """
    def create_widget(data, 
                      widgetType='label', 
                      layout=None, 
                      frame=None, 
                      row=None, 
                      col=None, 
                      labelText=None, 
                      entryWidth=None, 
                      xPad=None, 
                      yPad=None,
                      comboValues=None,
                      minVal = None,
                      maxVal = None,
                      valType = None,
                      val = None,
                      inval = None,
                      stickyVal = None,
                      cmd = None
                      ):
        
        #sets the default frame to be to root frame
        if frame is None:
              frame = data.rootFrame
              
        if widgetType == "tick":
            data.checkState= tk.IntVar()
        else:
            data.checkState = None
              
        #Functions required to create widgets stored in a dictionary               
        widgetDict = {
              "frame":tk.Frame,
              "labelframe":tk.LabelFrame,
              "label":tk.Label,
              "entry":tk.Entry,
              "combo":ttk.Combobox,
              "spin":ttk.Spinbox,
              "tick": tk.Checkbutton,
              "button": tk.Button
              }
        #Widget function called depending on function parameter "widgetType"
        #If an option is set as 'None' (i.e. left default), it is ignored
        data.Widget=widgetDict[widgetType](frame, text=labelText, width=entryWidth, values = comboValues, from_ = minVal, to=maxVal,
                                           validate=valType, validatecommand=val, invalidcommand = inval, variable=data.checkState,command=cmd)
        
        #Layout command stored in dictionary as string
        layoutDict = {
             "pack": 'data.Widget.pack(padx = xPad, pady = yPad)',
             "grid": 'data.Widget.grid(row=row, column=col, sticky=stickyVal, padx = xPad, pady = yPad)'
             }
        #exec() allows a string to be executed as if it were code
        if layout is not None:
            exec(layoutDict[layout])
            
        if widgetType=='tick':
            return data.Widget, data.checkState
        else:
            return data.Widget    
    
    """
    VALIDATION
    """
    """
    VALIDATION RULES
    """
    def validation(data,
                   widget, 
                   frame, 
                   valType,
                   minVal=None,
                   maxVal=None,
                   list=None ):
        widgetDict = {
            'spin':[data.validate_int, data.invalid_spin],
            'combo':[data.validate_combo, data.invalid_combo]
            }
        val, inval = widgetDict[widget][0], widgetDict[widget][1]   
        data.valid = (frame.register(val), '%P', minVal, maxVal, list)
        data.invalid = (frame.register(inval), '%P', minVal, maxVal, valType, list)
        return data.valid, data.invalid    
    """
    WIDGETS WITH VALIDATION
    """
    def valid_spinbox(data, row, col, min, max, labelText, frame, valType, width):
        min, max = min, max
        data.spinLabel = data.create_widget(widgetType='label', layout='grid', frame=frame, labelText=labelText, row=row, col=col)
        data.valid, data.invalid = data.validation(minVal=min, maxVal=max, valType=valType, frame=frame, widget='spin')
        return data.create_widget(widgetType='spin', layout='grid', row=row+1, col=col, frame=frame,
                                  entryWidth=width, minVal=min, maxVal=max, valType='focusout', val=data.valid, inval=data.invalid)    

    def valid_combobox(data, frame, labelText, row, col, list, width, valType):
        data.comboLabel = data.create_widget(widgetType='label', layout='grid', frame=frame, labelText=labelText, row=row, col=col)
        data.valid, data.invalid = data.validation(valType=valType, frame=frame, widget='combo', list=list)
        return data.create_widget(widgetType='combo', layout='grid', frame=frame, entryWidth=width, 
                                  comboValues=list, row=row+1, col=col, valType='focusout', val=data.valid, inval=data.invalid)
    
    """
    TYPE VALIDATION
    """
    #Checks to make sure that the value submitted is an integer, and within the correct range for the field.
    def validate_int(data, value, min, max, list):
        if value.isdigit() and int(value) in range(int(min),int(max)+1):
            return True
        else:
            return False
    
    #Checks to make sure entries are only text
    def validate_text(data, entry, valType):
        entryDict = {
            'Forename':data.forenameEntry.configure,
            'Surname':data.surnameEntry.configure
            }
        if entry.isalpha() == True:
            entryDict[valType](background='SystemWindow')
            return True
        else:
            entryDict[valType](background='#e06666')
            return False
    
    #Checks to make sure the value inputted is actually in the list
    def validate_combo(data, value, min, max, list):
        if value in list:
            return True
        return False
    
    """
    INVALID COMMANDS
    """
    def invalid_spin(data, value,  minVal, maxVal, valType, list):
        minVal=int(minVal)
        maxVal=int(maxVal)
        valuesDict = {
            'age': data.ageSpinbox.set,
            'courses': data.coursesSpinbox.set,
            'semesters': data.semestersSpinbox.set
            }
        try:
            float(value)
        except ValueError:
            valuesDict[valType](minVal)
            return
        value = float(value)
        if minVal <= value <= maxVal:
            valuesDict[valType](round(value))
        elif value < minVal:
            valuesDict[valType](minVal)
        else:
            valuesDict[valType](maxVal)

    def invalid_combo(data, value, min, max, valType, list ):
        comboDict = {
            'title':data.titleCombo.set,
            'nat':data.natCombo.set
            } 
        comboDict[valType]("")
        return
        
    
    """
    WIDGET SELECTION DETECTION
    """
    #Upon mouse click, the widget currently selected by the user is detected by the program. The labelframe
    #and the widget is then used to allow the user to press enter to update the fields. This is mainly used here
    #to ensure the numerical values in the spinboxes are of the correct value, without having to click away from
    #the field.
    def focus(data, event):
        global posData
        try:
            focussed = str(data.root.focus_get()).split(".")
            posData = [focussed.pop(2), focussed.pop(2)]
        except (IndexError, KeyError):
            return
        #print(posData)
        for idx, ele in enumerate(posData):
            posData[idx] = ele.replace('!',"")
    
    #When the user hits return, the data regarding which widget they are currently using is passed through to here.
    def return_check(data, event):
        #Checks which frame is selected and passes onto the function dedicated to widget shortcuts for that frame
        if posData[0].lower() == "labelframe":
            data.userinfo_shortcuts()
        elif posData[0].lower() == "labelframe2":
            data.courseframe_shortcuts()
        else:
            print("OTHER")
    
    """
    USERINFO SHORTCUTS
    """
    def userinfo_shortcuts(data):
        #Age Spinbox
        if posData[1].lower() == "spinbox":
            data.spinbox_shortcut(val=data.ageSpinbox, min=data.ageMin, max=data.ageMax, valType='age')
        
        #Comboboxes
        if posData[1].lower() == "combobox":
            data.combo_shortcut(data.titleCombo.get(), data.titleList, 'title')
        elif posData[1].lower()== "combobox2":
            data.combo_shortcut(data.natCombo.get(), data.natList, 'nat')
        
    """
    COURSEFRAME SHORTCUTS
    """
    def courseframe_shortcuts(data):
        #Courses Spinbox
        if posData[1].lower() == "spinbox":
            data.spinbox_shortcut(val=data.coursesSpinbox, min=data.coursesMin, max=data.coursesMax, valType='courses')
        elif posData[1].lower() == "spinbox2":
            data.spinbox_shortcut(val=data.semestersSpinbox, min=data.semestersMin, max=data.semestersMax, valType='semesters')
                    
    def spinbox_shortcut(data, val, min, max, valType):
        if data.validate_int(value=val.get(), min=min, max=max, list=None) == False:
            data.invalid_spin(value=val.get(), minVal=min, maxVal=max, valType=valType, list=None)
            
    def combo_shortcut(data, value, list, valType):
        if data.validate_combo(value=value, min=None, max=None, list=list) == False:
            data.invalid_combo(value=value, min=None, max=None, list=list, valType=valType)

    """
    CLEAR FORM
    """
    def clear(data, type):
        if type == 1:
            if messagebox.askyesno(title="Clear?", message="Are you sure you want to clear the form?") == False:
                return
        dataDict = {
            'Title':data.titleCombo,            
            'Age':data.ageSpinbox,
            'Nationality':data.natCombo,
            'No. completed courses':data.coursesSpinbox,
            'No. completed semesters':data.semestersSpinbox         
            }
        data.checkReg.set(0)
        data.checkTerms.set(0)
        for key in dataDict:
            dataDict[key].set("")
        data.forenameEntry.delete(0,tk.END)
        data.surnameEntry.delete(0,tk.END)
        return

    """
    ENTER DATA BUTTON
    """
    def enter_data(data):
        data.root.focus_get()
        dataDict = {
            'Title':data.titleCombo,            
            'Surname':data.surnameEntry,
            'Forename':data.forenameEntry,
            'Age':data.ageSpinbox,
            'Nationality':data.natCombo,
            'reg': data.checkReg,
            'No. completed courses':data.coursesSpinbox,
            'No. completed semesters':data.semestersSpinbox,
            'terms':data.checkTerms           
            }
        data.userData = []
        validEntry = True
        emptyValues = [False,'']
        for key in dataDict:
            value = data.get_data(dataDict[key])
            data.userData.append(value)
            if key.lower() == 'forename' or key.lower() == 'surname':
                if validEntry == False:
                    data.validate_text(dataDict[key].get(), key)
                elif validEntry == True:
                    validEntry = data.validate_text(dataDict[key].get(), key)
            if value == '':
                emptyValues = [True, key]
        if dataDict['terms'].get() == 0:
            data.error("Please accept the Terms and Conditions")
            return False 
        if validEntry == False or emptyValues[0] == True:
            dataDict['terms'].set(0)
            data.error(emptyValues[1]+" is Missing.")
            return False
        save = data.save_data(list=data.userData)
        if save == False:
            return False
        messagebox.askokcancel(title="Entry Successful", message="The entry has been added to your file succesfully.")
        data.clear(0)
        return True
        
    """
    GET DATA
    """
    def get_data(data,value):
            try:
                value.get()
            except AttributeError:
                return
            return value.get()

    """
    ERROR MESSAGE
    """
    def error(data, message):
        messagebox.showinfo(title="Error", message=message, icon='error')
        return

    """
    Open T&Cs
    """
    def open_terms(data):
        os.startfile("TermsandConditions.txt")
        return
    
    """
    Save to Excel
    """
    def save_data(data, list):
        if data.filepath == None:
            data.choose_filepath()
        data.new_workbook()
        #loads into a workbook from the currently selected filepath
        workbook = pyxl.load_workbook(data.filepath)
        sheet = workbook.active
        #generates an ID number based on the number of entries already in the worksheet
        list.insert(0, data.id(sheet))
        #converts any numbers in the list to integers so excel recognises them as numbers
        list = data.str_to_int(list)
        #converts binary values to "yes/no" labels
        list[6], list[9] = data.yes_no(list[6]), data.yes_no(list[9])
        #if the excel file is already open, an error will occur, so a message will appear indicating this
        try:
            sheet.append(list)                   
            workbook.save(data.filepath)
        except PermissionError:
            data.error("Please ensure the excel file is closed before trying to write to it.")
            return False
        return True
            
    """
    NEW EXCEL FILE
    """
    def new_workbook(data):
        #checks if the file exists already
        if not os.path.exists(data.filepath):
            #creates a new workbook
            workbook = pyxl.Workbook()
            #sets the currently selected sheet within the workbook as active
            sheet = workbook.active
            heading = ["ID", "Title", "Surname", "Forename", "Age", "Nationality", "Is Registered?", "No. Completed Courses",
                       "No. Completed Semesters", "Terms Accepted?"]
            #appends data from a list to the next available row in the excel sheet
            sheet.append(heading)
            #saves any changes made
            workbook.save(data.filepath)         
            return
        return
    
    """
    CHOOSE FILEPATH
    """
    def choose_filepath(data):
        if messagebox.askokcancel(title="Warning", 
                                  message="WARNING: Choosing a file containing data which was not previously stored there by this application may cause errors, and could cause irretrievable loss in data.",
                                  icon='warning'):   
            #creates a browse window to let the user choose where the file should be stored, automatically sets the extension
            #as .xlsx (excel)
            data.filepath = filedialog.asksaveasfilename(defaultextension='.xlsx')
            if data.filepath == '':
                data.filepath = None
            return
        return

    """
    ID GEN
    """
    def id(data, ws):
        maxRow = 0
        for i in range (1, ws.max_row+1):
            if not ws.cell(i,1).value:
                break
            maxRow+=1
        return maxRow
    
    """
    CONVERT STRING IN LIST TO INTEGER
    """
    def str_to_int(data, list):
        for i in range(len(list)):
            if type(list[i])==str:
                if list[i].isdigit():
                    list[i] = eval(list[i])
        return list

    """
    BINARY LABELS
    """
    def yes_no(data, value):
        if value == 0:
            return "No"
        if value == 1:
            return "Yes"

    """
    SETTING PAD VALUES
    """
    def pad_grid(data, frame, xPad, yPad):
        for widget in frame.winfo_children():
            widget.grid_configure(padx=xPad, pady=yPad)
    
    """
    TOOLBAR
    """
    """
    MENUS
    """
    def fileMenu(data):
        data.fileMenu = tk.Menu(data.menuBar, tearoff=0)
        data.fileMenu.add_command(label="Open Excel File", command=data.open_file)
        data.fileMenu.add_command(label="Change Filepath", command=data.choose_filepath)
        data.fileMenu.add_separator()
        data.fileMenu.add_command(label="Force Quit", command=lambda: data.quit(0))
        data.fileMenu.add_command(label="Quit", command=lambda: data.quit(1))
        data.menuBar.add_cascade(menu=data.fileMenu, label="File")
        return
    
    def helpMenu(data):
        data.helpMenu = tk.Menu(data.menuBar, tearoff=0)
        data.helpMenu.add_command(label="Terms & Conditions", command=data.open_terms)
        data.menuBar.add_cascade(menu=data.helpMenu, label="Help")
        return        

    def quit(data, type):
        if type == 0:
            sys.exit("Quitting Program...")
        else:
            if messagebox.askyesno(title="Quit?", message="Are you sure you want to quit?"):
                sys.exit("Quitting Program...")
                
    def open_file(data):
        if data.filepath==None:
            data.choose_filepath()
        os.startfile(data.filepath)
        return
        
        
    """
    AUTOFILL - FOR TESTING
    """
    def auto_fill(data, entry):
        name1 = tk.StringVar()
        name2 = tk.StringVar()
        data.forenameEntry.configure(textvariable=name1)
        name1.set("Fintan")
        data.surnameEntry.configure(textvariable=name2)
        name2.set("Spruce")
        data.titleCombo.set("Mr.")
        data.ageSpinbox.set(22)
        data.natCombo.set("White")
        data.checkReg.set(1)
        data.coursesSpinbox.set(4)
        data.semestersSpinbox.set(16)
        data.checkTerms.set(1)
        return
        
GUI = dataEntry()

